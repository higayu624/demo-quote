"""見積明細行の生成とExcel出力"""
from __future__ import annotations
import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")

COLUMNS = ["区分", "名称・仕様", "寸法/室名", "数量", "単位", "単価", "金額", "備考"]


def match_unit_price(name: str, prices: pd.DataFrame) -> tuple[float | None, str]:
    """名称に単価マスタのキーワードが含まれれば(単価, 単位)を返す"""
    if not isinstance(name, str):
        return None, ""
    for _, row in prices.iterrows():
        if row["キーワード"] in name:
            return float(row["単価"]), row["単位"]
    return None, ""


def _amount(qty, price):
    try:
        return round(float(qty) * float(price))
    except (TypeError, ValueError):
        return None


def rows_from_tategu(tategu: list[dict], prices: pd.DataFrame) -> list[dict]:
    """建具表の抽出結果 → 単価マスタ参照済みの見積明細行"""
    rows = []
    for it in tategu:
        name = f"{it.get('記号', '')} {it.get('種別', '')} {it.get('形式', '') or ''}".strip()
        qty = it.get("数量")
        price, unit = match_unit_price(f"{it.get('種別', '')}{it.get('形式', '') or ''}", prices)
        rows.append({
            "区分": "建具工事",
            "名称・仕様": name,
            "寸法/室名": f"W{it.get('W', '?')}×H{it.get('H', '?')}",
            "数量": qty,
            "単位": unit or "箇所",
            "単価": price,
            "金額": _amount(qty, price),
            "備考": it.get("備考", ""),
        })
    return rows


def rows_from_shiage(shiage: list[dict], prices: pd.DataFrame) -> list[dict]:
    """仕上表の抽出結果 → 単価マスタ参照済みの見積明細行"""
    rows = []
    for it in shiage:
        room = it.get("室名", "")
        area = it.get("床面積_m2")
        for part in ("床", "壁", "天井"):
            spec = it.get(part)
            if not spec:
                continue
            price, unit = match_unit_price(spec, prices)
            qty = area if (part == "床" and area) else None
            rows.append({
                "区分": "内装工事",
                "名称・仕様": f"{part}: {spec}",
                "寸法/室名": room,
                "数量": qty,
                "単位": unit or "m2",
                "単価": price,
                "金額": _amount(qty, price),
                "備考": it.get("備考", "") if part == "床" else "",
            })
    return rows


def build_estimate_from_rows(rows: list[dict], company: str = "御中") -> bytes:
    """明細行(表示テーブルそのまま)をExcelに書き出すだけ"""
    wb = Workbook()
    ws = wb.active
    ws.title = "見積内訳書"

    ws["A1"] = "見 積 内 訳 書"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"{company} 様"
    ws["F2"] = f"作成日: {date.today().isoformat()}"
    ws["A3"] = "※本書はAIによる自動拾い出し結果です。数量は必ずご確認ください。"
    ws["A3"].font = Font(size=9, color="C00000")

    headers = ["名称・仕様", "寸法/室名", "数量", "単位", "単価", "金額", "備考"]
    r = 5
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.fill = HEAD_FILL
        c.font = Font(bold=True)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    total = 0
    r += 1
    for kubun in ("建具工事", "内装工事", ""):
        group = [x for x in rows if (x.get("区分") or "") == kubun] if kubun else \
                [x for x in rows if (x.get("区分") or "") not in ("建具工事", "内装工事")]
        if not group:
            continue
        if kubun:
            ws.cell(row=r, column=1, value=f"【{kubun}】").font = Font(bold=True)
            r += 1
        for x in group:
            amount = _amount(x.get("数量"), x.get("単価"))
            vals = [x.get("名称・仕様"), x.get("寸法/室名"), x.get("数量"),
                    x.get("単位"), x.get("単価"), amount, x.get("備考")]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v)
                c.border = BORDER
                if col in (5, 6) and v is not None:
                    c.number_format = "#,##0"
            if amount:
                total += amount
            r += 1

    r += 1
    ws.cell(row=r, column=5, value="小計(参考)").font = Font(bold=True)
    c = ws.cell(row=r, column=6, value=total)
    c.font = Font(bold=True)
    c.number_format = "#,##0"

    widths = [34, 18, 8, 6, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
